import csv
import io
import json
import re
import unicodedata
from datetime import datetime, timezone
from hashlib import sha256

from openpyxl import load_workbook

from ..extensions import db
from ..models import Company, CompanyActivity, Contact, Opportunity, Project
from .entity_resolution import normalize_domain, normalize_name, resolve_company, resolve_project

MAX_IMPORT_ROWS = 5000

FIELD_ALIASES = {
    "company": ["empresa", "company", "razao social", "razon social", "nome empresa", "nombre empresa", "cliente"],
    "legal_name": ["razao social", "razon social", "legal name", "nome legal", "nombre legal"],
    "ruc": ["ruc", "cnpj", "registro", "registration id", "tax id"],
    "website": ["site", "website", "sitio", "web", "url", "pagina web", "página web"],
    "sector": ["setor", "sector", "segmento", "industria", "indústria"],
    "city": ["cidade", "ciudad", "city", "localidade", "localidad"],
    "department": ["departamento", "estado", "region", "regiao", "región", "uf"],
    "country": ["pais", "país", "country"],
    "company_email": ["email empresa", "correo empresa", "email corporativo", "correo corporativo", "email geral", "correo general"],
    "company_phone": ["telefone empresa", "telefono empresa", "tel empresa", "phone company", "telefone geral", "telefono general"],
    "company_whatsapp": ["whatsapp empresa", "whats empresa", "wa empresa"],
    "contact_name": ["contato", "contacto", "contact", "nome contato", "nombre contacto", "responsavel", "responsable"],
    "contact_role": ["cargo", "funcao", "função", "puesto", "role", "area", "área", "departamento contato"],
    "contact_email": ["email contato", "correo contacto", "email", "correo", "e-mail"],
    "contact_phone": ["telefone", "telefono", "phone", "celular", "movil", "móvil"],
    "contact_whatsapp": ["whatsapp", "whats", "wa"],
    "date": ["data", "fecha", "date", "data contato", "fecha contacto"],
    "channel": ["canal", "channel", "meio", "medio"],
    "status": ["status", "estado crm", "etapa", "stage", "situacao", "situação"],
    "observation": ["observacao", "observação", "observaciones", "observacion", "nota", "notas", "comentario", "comentários", "resumo", "historico", "histórico"],
    "next_action": ["proxima acao", "próxima ação", "proxima accion", "próxima acción", "next action", "follow up", "follow-up"],
    "next_action_at": ["data follow up", "fecha follow up", "proxima data", "próxima data", "next action date"],
    "owner": ["responsavel comercial", "responsable comercial", "owner", "vendedor", "comercial"],
}

STATUS_MAP = {
    "novo": "NOVO", "nuevo": "NOVO",
    "qualificado": "QUALIFICADO", "calificado": "QUALIFICADO", "qualified": "QUALIFICADO",
    "contato realizado": "CONTATO_REALIZADO", "contacto realizado": "CONTATO_REALIZADO", "contactado": "CONTATO_REALIZADO", "contacted": "CONTATO_REALIZADO",
    "respondeu": "RESPONDEU", "respondio": "RESPONDEU", "respondió": "RESPONDEU", "resposta": "RESPONDEU", "reply": "RESPONDEU",
    "visita": "VISITA", "visita marcada": "VISITA", "visita agendada": "VISITA", "reuniao": "VISITA", "reunião": "VISITA", "reunion": "VISITA", "reunión": "VISITA",
    "orcamento": "ORCAMENTO", "orçamento": "ORCAMENTO", "presupuesto": "ORCAMENTO", "proposta": "ORCAMENTO", "propuesta": "ORCAMENTO",
    "negociacao": "NEGOCIACAO", "negociação": "NEGOCIACAO", "negociacion": "NEGOCIACAO", "negociación": "NEGOCIACAO",
    "ganho": "GANHO", "ganado": "GANHO", "won": "GANHO",
    "perdido": "PERDIDO", "lost": "PERDIDO",
    "monitoramento": "MONITORAMENTO", "seguimiento": "MONITORAMENTO", "follow up": "MONITORAMENTO", "follow-up": "MONITORAMENTO",
    "descartado": "DESCARTADO", "descartada": "DESCARTADO",
}
STATUS_RANK = {"NOVO": 0, "QUALIFICADO": 1, "CONTATO_REALIZADO": 2, "RESPONDEU": 3, "VISITA": 4, "ORCAMENTO": 5, "NEGOCIACAO": 6, "GANHO": 7}

ACTIVITY_MAP = {
    "email": "EMAIL_SENT", "e-mail": "EMAIL_SENT", "correo": "EMAIL_SENT", "correio": "EMAIL_SENT",
    "whatsapp": "WHATSAPP_SENT", "whats": "WHATSAPP_SENT", "wa": "WHATSAPP_SENT",
    "ligacao": "CALL", "ligação": "CALL", "llamada": "CALL", "call": "CALL", "telefone": "CALL", "telefono": "CALL",
    "resposta": "REPLY", "respuesta": "REPLY", "reply": "REPLY", "respondido": "REPLY",
    "reuniao": "MEETING", "reunião": "MEETING", "reunion": "MEETING", "reunión": "MEETING", "meeting": "MEETING",
    "visita": "VISIT", "visit": "VISIT",
    "proposta": "PROPOSAL_SENT", "propuesta": "PROPOSAL_SENT", "presupuesto": "PROPOSAL_SENT", "orcamento": "PROPOSAL_SENT", "orçamento": "PROPOSAL_SENT",
    "follow up": "FOLLOW_UP", "follow-up": "FOLLOW_UP", "seguimiento": "FOLLOW_UP", "acompanhamento": "FOLLOW_UP",
}


def _norm(value):
    value = unicodedata.normalize("NFKD", str(value or "").casefold())
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _phone(value):
    return re.sub(r"\D", "", _text(value))


def _parse_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    raw = _text(value)
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _read_rows(upload):
    name = (upload.filename or "").lower()
    raw = upload.read()
    if len(raw) > 12 * 1024 * 1024:
        raise ValueError("La planilla supera el límite de 12 MB")
    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        try:
            headers = [_text(v) or f"Columna {idx+1}" for idx, v in enumerate(next(iterator))]
        except StopIteration:
            return [], []
        rows = []
        for values in iterator:
            row = {headers[i]: _text(values[i]) if i < len(values) else "" for i in range(len(headers))}
            if any(row.values()):
                rows.append(row)
            if len(rows) >= MAX_IMPORT_ROWS:
                break
        return headers, rows
    if not name.endswith(".csv"):
        raise ValueError("Formato no soportado. Use .xlsx o .csv")
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("No se pudo leer el archivo CSV")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [str(h or "").strip() for h in (reader.fieldnames or [])]
    rows = []
    for row in reader:
        clean = {str(k or "").strip(): _text(v) for k, v in row.items()}
        if any(clean.values()):
            rows.append(clean)
        if len(rows) >= MAX_IMPORT_ROWS:
            break
    return headers, rows


def detect_mapping(headers):
    normalized = {_norm(h): h for h in headers}
    result = {}
    used = set()
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            key = _norm(alias)
            exact = normalized.get(key)
            if exact and exact not in used:
                result[field] = exact
                used.add(exact)
                break
        if field in result:
            continue
        for norm_header, original in normalized.items():
            if original in used:
                continue
            if any(_norm(alias) in norm_header or norm_header in _norm(alias) for alias in aliases if len(_norm(alias)) >= 4):
                result[field] = original
                used.add(original)
                break
    return result


def preview(upload):
    headers, rows = _read_rows(upload)
    return {
        "columns": headers,
        "detectedMapping": detect_mapping(headers),
        "rowCount": len(rows),
        "sample": rows[:8],
        "truncated": len(rows) >= MAX_IMPORT_ROWS,
    }


def _value(row, mapping, key):
    column = mapping.get(key)
    return _text(row.get(column)) if column else ""


def _contact_for_row(tenant_id, company, data):
    name = data.get("contact_name") or ""
    role = data.get("contact_role") or ""
    email = data.get("contact_email") or ""
    phone = data.get("contact_phone") or ""
    whatsapp = data.get("contact_whatsapp") or ""
    if not any((name, role, email, phone, whatsapp)):
        return None, False, False
    existing = Contact.query.filter_by(tenant_id=tenant_id, company_id=company.id, status="ACTIVE").all()
    email_key = email.casefold().strip()
    phone_key = _phone(whatsapp or phone)
    name_key = normalize_name(name)
    row = None
    for item in existing:
        if email_key and (item.email or "").casefold().strip() == email_key:
            row = item; break
        if phone_key and phone_key in {_phone(item.whatsapp), _phone(item.phone)}:
            row = item; break
        if name_key and normalize_name(item.name) == name_key and (not role or _norm(item.role) == _norm(role)):
            row = item; break
    created = False
    updated = False
    if not row:
        display_name = name or role or email or whatsapp or phone or "Contacto importado"
        row = Contact(tenant_id=tenant_id, company_id=company.id, name=display_name, role=role or None, email=email or None,
                      phone=phone or None, whatsapp=whatsapp or None, source_url="HISTORICO_IMPORTADO", confidence=75, status="ACTIVE")
        db.session.add(row)
        db.session.flush()
        created = True
    else:
        for attr, value in (("name", name), ("role", role), ("email", email), ("phone", phone), ("whatsapp", whatsapp)):
            if value and not getattr(row, attr):
                setattr(row, attr, value)
                updated = True
    return row, created, updated


def _status(value):
    norm = _norm(value)
    if not norm:
        return None
    if value and str(value).strip().upper() in (set(STATUS_RANK) | {"PERDIDO", "MONITORAMENTO", "DESCARTADO"}):
        return str(value).strip().upper()
    return STATUS_MAP.get(norm)


def _activity_type(channel, status, observation):
    blob = " ".join((_norm(channel), _norm(status), _norm(observation)))
    for key, value in ACTIVITY_MAP.items():
        if _norm(key) and _norm(key) in blob:
            return value
    mapped_status = _status(status)
    return {"CONTATO_REALIZADO": "FOLLOW_UP", "RESPONDEU": "REPLY", "VISITA": "MEETING", "ORCAMENTO": "PROPOSAL_SENT", "MONITORAMENTO": "FOLLOW_UP"}.get(mapped_status)


def _ensure_opportunity(tenant_id, company, data, owner_name):
    opportunity = Opportunity.query.join(Project).filter(Project.company_id == company.id, Opportunity.tenant_id == tenant_id).order_by(Opportunity.updated_at.desc()).first()
    wanted = _status(data.get("status"))
    created = False
    if not opportunity:
        project = resolve_project(tenant_id, company, "Histórico comercial importado", city=company.city or "Por validar", department=company.department or "Por validar", country=company.country, project_type="IMPORTED_HISTORY", stage="Histórico importado")
        opportunity = Opportunity(tenant_id=tenant_id, project=project, event_type="IMPORTED_HISTORY", score=50, level="MEDIUM",
                                  status=wanted or "QUALIFICADO", products=[], evidence="Registro incorporado desde planilla/histórico comercial.",
                                  source_name="Importación de histórico", probability=20, owner_name=owner_name or "Equipo comercial")
        db.session.add(opportunity)
        db.session.flush()
        created = True
    elif wanted:
        current = opportunity.status or "NOVO"
        if wanted in {"PERDIDO", "DESCARTADO", "GANHO"} or STATUS_RANK.get(wanted, -1) > STATUS_RANK.get(current, -1):
            opportunity.status = wanted
    if data.get("next_action_at"):
        dt = _parse_datetime(data.get("next_action_at"))
        if dt:
            opportunity.next_action_at = dt
    if owner_name:
        opportunity.owner_name = owner_name
    return opportunity, created


def _activity_exists(tenant_id, company_id, activity_type, occurred_at, summary):
    q = CompanyActivity.query.filter_by(tenant_id=tenant_id, company_id=company_id, activity_type=activity_type)
    if occurred_at:
        start = occurred_at.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start.replace(hour=23, minute=59, second=59, microsecond=999999)
        q = q.filter(CompanyActivity.occurred_at.between(start, end))
    for row in q.limit(50).all():
        if _norm(row.summary) == _norm(summary):
            return True
    return False


def import_rows(upload, mapping, tenant, user):
    headers, rows = _read_rows(upload)
    if not mapping:
        mapping = detect_mapping(headers)
    if not mapping.get("company"):
        raise ValueError("Seleccione la columna que contiene el nombre de la empresa")
    stats = {"rows": len(rows), "companiesCreated": 0, "companiesUpdated": 0, "contactsCreated": 0, "contactsUpdated": 0,
             "opportunitiesCreated": 0, "activitiesCreated": 0, "duplicatesSkipped": 0, "errors": []}
    seen_companies = set()
    for idx, row in enumerate(rows, start=2):
        try:
            data = {key: _value(row, mapping, key) for key in FIELD_ALIASES}
            company_name = data.get("company")
            if not company_name:
                stats["duplicatesSkipped"] += 1
                continue
            country = data.get("country") or (tenant.settings or {}).get("default_country") or ("Brasil" if (tenant.settings or {}).get("language") == "pt-BR" else "Paraguay")
            ruc = data.get("ruc") or None
            domain = normalize_domain(data.get("website"))
            normalized_company = normalize_name(company_name)
            existing_company = None
            base = Company.query.filter_by(tenant_id=tenant.id, status="ACTIVE")
            if ruc:
                existing_company = base.filter((Company.ruc == ruc) | (Company.registration_id == ruc)).first()
            if not existing_company and domain:
                existing_company = base.filter_by(domain=domain).first()
            if not existing_company and normalized_company:
                existing_company = base.filter_by(normalized_name=normalized_company, country=country).first()
            company = resolve_company(tenant.id, company_name, website=data.get("website") or None, registration_id=ruc,
                                      ruc=ruc, legal_name=data.get("legal_name") or None, sector=data.get("sector") or None,
                                      city=data.get("city") or None, department=data.get("department") or None, country=country,
                                      email_business=data.get("company_email") or None, phone_business=data.get("company_phone") or None,
                                      whatsapp=data.get("company_whatsapp") or None)
            was_seen = company.id in seen_companies
            if not was_seen:
                stats["companiesUpdated" if existing_company else "companiesCreated"] += 1
                seen_companies.add(company.id)
            if ruc and not company.ruc:
                company.ruc = ruc
            if data.get("legal_name") and not company.legal_name:
                company.legal_name = data["legal_name"]
            contact, c_created, c_updated = _contact_for_row(tenant.id, company, data)
            stats["contactsCreated"] += int(c_created)
            stats["contactsUpdated"] += int(c_updated)
            opportunity, o_created = _ensure_opportunity(tenant.id, company, data, data.get("owner") or (user.name if user else "Equipo comercial"))
            stats["opportunitiesCreated"] += int(o_created)
            activity_type = _activity_type(data.get("channel"), data.get("status"), data.get("observation"))
            if activity_type:
                occurred = _parse_datetime(data.get("date")) or datetime.now(timezone.utc)
                summary = data.get("observation") or f"Actividad importada desde histórico comercial · {data.get('channel') or data.get('status') or activity_type}"
                if _activity_exists(tenant.id, company.id, activity_type, occurred, summary):
                    stats["duplicatesSkipped"] += 1
                else:
                    channel = data.get("channel") or {"EMAIL_SENT": "EMAIL", "WHATSAPP_SENT": "WHATSAPP", "CALL": "CALL", "MEETING": "PRESENCIAL", "VISIT": "PRESENCIAL"}.get(activity_type)
                    db.session.add(CompanyActivity(tenant_id=tenant.id, company_id=company.id, opportunity_id=opportunity.id,
                        contact_id=contact.id if contact else None, activity_type=activity_type, channel=(channel or "HISTORICO_IMPORTADO")[:40],
                        direction="INBOUND" if activity_type == "REPLY" else "OUTBOUND", subject="Histórico comercial importado",
                        summary=summary, outcome="IMPORTED", next_action=data.get("next_action") or None,
                        next_action_at=_parse_datetime(data.get("next_action_at")), occurred_at=occurred,
                        created_by=data.get("owner") or (user.name if user else "Equipo comercial"),
                        extra_data={"source": "HISTORICO_IMPORTADO", "row": idx}))
                    stats["activitiesCreated"] += 1
        except Exception as exc:
            stats["errors"].append({"row": idx, "error": str(exc)[:240]})
            if len(stats["errors"]) >= 50:
                break
    return stats
